"""
apply_absences.py  Reads "Absences" sheet from weekly_overrides.xlsx and
inserts rows into the team_absences SQLite table.

One row per person per absent day. Running twice produces the same result
(INSERT OR REPLACE + UNIQUE constraint on week_start, member, date).

Usage (CLI):
    python -m scripts.etl.apply_absences                     # all weeks
    python -m scripts.etl.apply_absences --week 2026-03-20  # one week
    python -m scripts.etl.apply_absences --dry-run           # preview only
"""
from __future__ import annotations

import argparse
import logging
import sqlite3
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)

# Row 1 = headers, data starts at row 2 (1-indexed, openpyxl convention)
_DATA_START_ROW = 2

# Column indices (1-indexed)
_COL_WEEK_START = 1   # A
_COL_MEMBER     = 2   # B
_COL_DATE       = 3   # C
_COL_TYPE       = 4   # D

_VALID_TYPES = frozenset({"Holiday", "Sickness", "Other", "Half Day"})


# ---------------------------------------------------------------------------
# Schema
# ---------------------------------------------------------------------------

def ensure_schema(conn: sqlite3.Connection) -> None:
    """Create team_absences table if it doesn't exist."""
    conn.execute("""
        CREATE TABLE IF NOT EXISTS team_absences (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            week_start TEXT NOT NULL,
            member     TEXT NOT NULL,
            date       TEXT NOT NULL,
            type       TEXT NOT NULL,
            UNIQUE(week_start, member, date)
        )
    """)
    columns = {row[1] for row in conn.execute("PRAGMA table_info(team_absences)").fetchall()}
    if "source" not in columns:
        conn.execute("ALTER TABLE team_absences ADD COLUMN source TEXT NOT NULL DEFAULT 'weekly_overrides'")
    conn.execute("""
        CREATE TABLE IF NOT EXISTS team_absence_deletions (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            member      TEXT NOT NULL,
            date        TEXT NOT NULL,
            source      TEXT NOT NULL DEFAULT 'calendar',
            deleted_at  TEXT NOT NULL,
            created_by  TEXT,
            UNIQUE(member, date)
        )
    """)
    conn.commit()


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _parse_date(val: Any) -> str:
    """Parse dd/mm/yyyy or YYYY-MM-DD (or datetime object)  YYYY-MM-DD."""
    if not val:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, date):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, (int, float)):
        return (datetime(1899, 12, 30) + timedelta(days=float(val))).strftime("%Y-%m-%d")

    text = str(val).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return text


def _cell(row: tuple[Any, ...], col: int) -> str:
    """Return stripped string value for 1-indexed column, or '' if absent."""
    idx = col - 1
    if idx >= len(row):
        return ""
    val = row[idx]
    if val is None:
        return ""
    if isinstance(val, (datetime, date)):
        return _parse_date(val)
    return str(val).strip()


def _read_absences_file(path: Path) -> list[dict[str, str]]:
    """
    Read the Absences sheet from weekly_overrides.xlsx.
    Raises ValueError if the Absences sheet is not found.
    Raises FileNotFoundError / PermissionError  caller handles gracefully.
    """
    try:
        import openpyxl  # noqa: PLC0415
    except ImportError as exc:  # pragma: no cover
        raise ImportError("openpyxl is required: pip install openpyxl") from exc

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if "Absences" not in wb.sheetnames:
            raise ValueError(
                f"Sheet 'Absences' not found in {path}. "
                f"Available sheets: {wb.sheetnames}"
            )
        ws = wb["Absences"]
        rows: list[dict[str, str]] = []
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row_idx < _DATA_START_ROW:
                continue
            if all(v is None or str(v).strip() == "" for v in row):
                continue
            rows.append({
                "week_start": _parse_date(_cell(row, _COL_WEEK_START)),
                "member":     _cell(row, _COL_MEMBER),
                "date":       _parse_date(_cell(row, _COL_DATE)),
                "type":       _cell(row, _COL_TYPE),
            })
    finally:
        wb.close()
    return rows


def _validate_row(
    raw: dict[str, str],
    team_members: list[str],
    row_num: int,
) -> dict[str, str] | None:
    """Validate one raw row. Returns cleaned dict or None (logs warning)."""
    errors: list[str] = []

    if not raw["week_start"]:
        errors.append("week_start is empty or unparseable")
    if not raw["member"]:
        errors.append("member is empty")
    elif raw["member"] not in team_members:
        errors.append(f"member {raw['member']!r} not in TEAM_MEMBERS {team_members}")
    if not raw["date"]:
        errors.append("date is empty or unparseable")
    if raw["type"] not in _VALID_TYPES:
        errors.append(f"type {raw['type']!r} must be one of {sorted(_VALID_TYPES)}")

    if errors:
        logger.warning("Row %d skipped: %s", row_num, "; ".join(errors))
        return None

    return {
        "week_start": raw["week_start"],
        "member":     raw["member"],
        "date":       raw["date"],
        "type":       raw["type"],
    }


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def apply_absences(
    db_path: Path | None = None,
    overrides_path: Path | None = None,
    week: str | None = None,
    dry_run: bool = False,
    _conn: sqlite3.Connection | None = None,  # injected for unit tests
) -> dict[str, Any]:
    """
    Apply absences from the Absences sheet of weekly_overrides.xlsx.

    Parameters
    ----------
    db_path:        Path to invoices.db. Defaults to scripts.paths.DB_PATH.
    overrides_path: Path to weekly_overrides.xlsx. Defaults to DATA_DIR / "weekly_overrides.xlsx".
    week:           If given, only process rows whose week_start matches (YYYY-MM-DD).
    dry_run:        If True, print planned inserts without executing them.
    _conn:          Injected connection (for unit tests  skips db_path entirely).

    Returns
    -------
    dict with keys: applied, rows_inserted, skipped, errors
    """
    from scripts.paths import DB_PATH, DATA_DIR  # noqa: PLC0415
    from scripts.config import ALL_MEMBERS       # noqa: PLC0415  (incl. former members for historical rows)

    if db_path is None:
        db_path = DB_PATH
    if overrides_path is None:
        overrides_path = DATA_DIR / "weekly_overrides.xlsx"

    result: dict[str, Any] = {
        "applied": 0,
        "rows_inserted": 0,
        "skipped": 0,
        "errors": [],
    }

    # --- Graceful degradation: missing file ---
    if not overrides_path.exists():
        logger.info("No overrides file at %s  skipping absences.", overrides_path)
        return result

    # --- Read Excel ---
    try:
        raw_rows = _read_absences_file(overrides_path)
    except ValueError as exc:
        # Sheet missing is expected when TL hasn't added it yet
        logger.info("Absences sheet not found: %s", exc)
        return result
    except PermissionError:
        logger.warning("Overrides file is locked (PermissionError): %s", overrides_path)
        return result
    except Exception as exc:  # noqa: BLE001
        logger.error("Failed to read absences file: %s", exc)
        result["errors"].append(str(exc))
        return result

    if not raw_rows:
        logger.info("Absences sheet is empty.")
        return result

    # --- Validate rows ---
    validated: list[dict[str, str]] = []
    for idx, raw in enumerate(raw_rows, start=_DATA_START_ROW):
        if week and raw["week_start"] and raw["week_start"] != week:
            continue
        cleaned = _validate_row(raw, ALL_MEMBERS, idx)
        if cleaned is None:
            result["skipped"] += 1
        else:
            validated.append(cleaned)

    if not validated:
        logger.info("No valid absence rows (week filter: %s).", week)
        return result

    # --- Apply / dry-run ---
    own_conn = _conn is None
    conn: sqlite3.Connection = _conn or sqlite3.connect(str(db_path))
    try:
        ensure_schema(conn)

        for row in validated:
            if dry_run:
                print(
                    f"[DRY-RUN] Would insert: week={row['week_start']} "
                    f"member={row['member']} date={row['date']} type={row['type']}"
                )
                result["applied"] += 1
                continue

            try:
                conn.execute(
                    "INSERT OR REPLACE INTO team_absences (week_start, member, date, type, source) "
                    "VALUES (?, ?, ?, ?, ?)",
                    (row["week_start"], row["member"], row["date"], row["type"], "weekly_overrides"),
                )
                result["applied"] += 1
                result["rows_inserted"] += 1
            except sqlite3.Error as exc:
                msg = (
                    f"DB error: week={row['week_start']} "
                    f"member={row['member']}: {exc}"
                )
                logger.error(msg)
                result["errors"].append(msg)

        if not dry_run:
            conn.commit()
            logger.info(
                "Absences applied: %d row(s) inserted for week=%s.",
                result["rows_inserted"],
                week or "all",
            )
    finally:
        if own_conn:
            conn.close()

    return result


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def _build_arg_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description="Apply absences from the Absences sheet of weekly_overrides.xlsx."
    )
    p.add_argument("--week", metavar="YYYY-MM-DD", default=None,
                   help="Only process rows for this week_start date.")
    p.add_argument("--dry-run", action="store_true", default=False,
                   help="Print planned inserts without executing.")
    p.add_argument("--db", metavar="PATH", default=None,
                   help="Path to invoices.db (default: scripts.paths.DB_PATH).")
    p.add_argument("--overrides", metavar="PATH", default=None,
                   help="Path to weekly_overrides.xlsx (default: data/weekly_overrides.xlsx).")
    return p


if __name__ == "__main__":
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)-8s %(name)s  %(message)s",
    )
    parser = _build_arg_parser()
    args = parser.parse_args()
    summary = apply_absences(
        db_path=Path(args.db) if args.db else None,
        overrides_path=Path(args.overrides) if args.overrides else None,
        week=args.week,
        dry_run=args.dry_run,
    )
    print(
        f"Done. applied={summary['applied']}  "
        f"rows_inserted={summary['rows_inserted']}  "
        f"skipped={summary['skipped']}  "
        f"errors={len(summary['errors'])}"
    )
    if summary["errors"]:
        for err in summary["errors"]:
            print(f"  ERROR: {err}")
