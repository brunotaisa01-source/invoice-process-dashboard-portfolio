"""
apply_overrides.py - Apply team-member reassignments from weekly_overrides.xlsx.

Reads the Overrides sheet, validates each row, then issues UPDATE statements
against the invoices table.  Preserves the original team_member in the
`original_team_member` column (set once; never overwritten on subsequent runs).

Usage (CLI):
    python -m scripts.etl.apply_overrides                     # all weeks
    python -m scripts.etl.apply_overrides --week 2026-03-21  # one week
    python -m scripts.etl.apply_overrides --dry-run           # preview only
"""
from __future__ import annotations

import argparse
import logging
import sqlite3
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Day abbreviation -> SQLite strftime('%w') value (Sun=0 ... Sat=6)
# ---------------------------------------------------------------------------
_DAY_TO_STRFTIME: dict[str, str] = {
    "Mon": "1",
    "Tue": "2",
    "Wed": "3",
    "Thu": "4",
    "Fri": "5",
}
_VALID_DAYS = frozenset(_DAY_TO_STRFTIME)

# Row in the sheet where real data begins (1-indexed, openpyxl convention)
_DATA_START_ROW = 9

# Column indices (1-indexed)
_COL_WEEK_START     = 1   # A
_COL_DAY            = 2   # B
_COL_COMPANY_CODE   = 3   # C
_COL_SCOPE          = 4   # D
_COL_SUPPLIER_NUMBER= 5   # E
_COL_FROM_MEMBER    = 6   # F
_COL_TO_MEMBER      = 7   # G
_COL_INVOICE_COUNT  = 8   # H  (used as LIMIT in SQL when present)
_COL_DAYS           = 9   # I  (informational  not used in SQL)
_COL_NOTES          = 10  # J  (informational  not used in SQL)


# ---------------------------------------------------------------------------
# Schema migration
# ---------------------------------------------------------------------------

def migrate_schema(conn: sqlite3.Connection) -> None:
    """Add original_team_member column if missing (backward-compatible)."""
    cursor = conn.execute("PRAGMA table_info(invoices)")
    columns = {row[1] for row in cursor.fetchall()}
    if "original_team_member" not in columns:
        conn.execute(
            "ALTER TABLE invoices ADD COLUMN original_team_member TEXT"
        )
        conn.commit()
        logger.info("Schema migrated: original_team_member column added.")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _parse_date(val: Any) -> str:
    """Parse a date value and return YYYY-MM-DD.

    Accepts text, Excel date objects, and Excel serial date numbers.
    """
    if not val:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, date):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, (int, float)):
        return (datetime(1899, 12, 30) + timedelta(days=float(val))).strftime("%Y-%m-%d")

    text = str(val).strip()
    # Try dd/mm/yyyy first (TL format)
    for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return text  # return as-is if unparseable (validation will catch it)


def _cell(row: tuple[Any, ...], col: int) -> str:
    """Return stripped string value for 1-indexed column, or '' if absent.

    Handles datetime objects returned by openpyxl for date-formatted cells,
    converting them to YYYY-MM-DD before further parsing.
    """
    idx = col - 1
    if idx >= len(row):
        return ""
    val = row[idx]
    if val is None:
        return ""
    if isinstance(val, (datetime, date)):
        return _parse_date(val)
    return str(val).strip()


def _parse_days(day_str: str) -> list[str] | None:
    """
    Parse the 'day' cell.

    Returns:
        None          FULL WEEK (no day filter)
        list[str]     list of strftime('%w') digits for named days
        Raises ValueError if the string is unrecognisable.
    """
    if day_str.upper() == "FULL WEEK":
        return None

    parts = [p.strip() for p in day_str.split(",") if p.strip()]
    if not parts:
        raise ValueError(f"Empty day value: {day_str!r}")

    unknown = [p for p in parts if p not in _VALID_DAYS]
    if unknown:
        raise ValueError(
            f"Unknown day abbreviation(s): {unknown}. "
            f"Valid: {sorted(_VALID_DAYS)}"
        )
    return [_DAY_TO_STRFTIME[p] for p in parts]


def _read_overrides_file(
    overrides_path: Path,
) -> list[dict[str, str]]:
    """
    Open the Excel file and return a list of raw row dicts (strings only).
    Skips completely blank rows.
    Raises FileNotFoundError / PermissionError  caller handles gracefully.
    """
    try:
        import openpyxl  # noqa: PLC0415
    except ImportError as exc:  # pragma: no cover
        raise ImportError(
            "openpyxl is required: pip install openpyxl"
        ) from exc

    wb = openpyxl.load_workbook(overrides_path, read_only=True, data_only=True)
    try:
        if "Overrides" not in wb.sheetnames:
            raise ValueError(
                f"Sheet 'Overrides' not found in {overrides_path}. "
                f"Available sheets: {wb.sheetnames}"
            )
        ws = wb["Overrides"]

        rows: list[dict[str, str]] = []
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row_idx < _DATA_START_ROW:
                continue
            # Skip blank rows (all cells None or empty string)
            if all((v is None or str(v).strip() == "") for v in row):
                continue
            # Parse invoice_count: keep as int or None
            raw_count = row[_COL_INVOICE_COUNT - 1] if len(row) >= _COL_INVOICE_COUNT else None
            inv_count: int | None = None
            if raw_count is not None and str(raw_count).strip():
                try:
                    inv_count = int(float(str(raw_count).strip()))
                except (ValueError, TypeError):
                    inv_count = None

            rows.append({
                "week_start":      _parse_date(_cell(row, _COL_WEEK_START)),
                "day":             _cell(row, _COL_DAY),
                "company_code":    _cell(row, _COL_COMPANY_CODE),
                "scope":           _cell(row, _COL_SCOPE),
                "supplier_number": _cell(row, _COL_SUPPLIER_NUMBER),
                "from_member":     _cell(row, _COL_FROM_MEMBER),
                "to_member":       _cell(row, _COL_TO_MEMBER),
                "invoice_count":   inv_count,
            })
    finally:
        wb.close()
    return rows


# ---------------------------------------------------------------------------
# Validation
# ---------------------------------------------------------------------------

def _validate_row(
    raw: dict[str, str],
    team_members: list[str],
    company_code_country_map: dict[str, str],
    row_num: int,
) -> dict[str, Any] | None:
    """
    Validate one raw row.  Returns cleaned dict or None (with warning logged).
    """
    errors: list[str] = []

    week_start = raw["week_start"]
    if not week_start:
        errors.append("week_start is empty")

    to_member = raw["to_member"]
    if not to_member:
        errors.append("to_member is empty")
    elif to_member not in team_members:
        errors.append(
            f"to_member {to_member!r} not in TEAM_MEMBERS {team_members}"
        )

    company_code = raw["company_code"]
    if not company_code:
        errors.append("company_code is empty")
    elif company_code not in company_code_country_map:
        errors.append(
            f"company_code {company_code!r} not in COMPANY_CODE_COUNTRY_MAP"
        )

    scope = raw["scope"].upper() if raw["scope"] else ""
    if scope not in ("ALL", "SUPPLIER"):
        errors.append(f"scope must be 'ALL' or 'SUPPLIER', got {raw['scope']!r}")

    supplier_number = raw["supplier_number"]
    if scope == "SUPPLIER" and not supplier_number:
        errors.append("scope=SUPPLIER requires a non-empty supplier_number")

    day_str = raw["day"]
    day_digits: list[str] | None = None
    if not day_str:
        errors.append("day is empty (expected 'FULL WEEK' or day list)")
    else:
        try:
            day_digits = _parse_days(day_str)
        except ValueError as exc:
            errors.append(str(exc))

    from_member = raw["from_member"]
    if from_member and from_member not in team_members:
        errors.append(
            f"from_member {from_member!r} not in TEAM_MEMBERS {team_members}"
        )

    if errors:
        logger.warning(
            "Row %d skipped  validation failed: %s", row_num, "; ".join(errors)
        )
        return None

    return {
        "week_start":      week_start,
        "day_digits":      day_digits,       # None = FULL WEEK
        "company_code":    company_code,
        "scope":           scope,
        "supplier_number": supplier_number,
        "from_member":     from_member,
        "to_member":       to_member,
        "invoice_count":   raw.get("invoice_count"),  # int or None
    }


# ---------------------------------------------------------------------------
# SQL builder
# ---------------------------------------------------------------------------

def _build_update(row: dict[str, Any]) -> tuple[str, list[Any]]:
    """
    Build (sql, params) for one validated override row.

    The COALESCE guard ensures original_team_member is only set on the first
    override; subsequent runs preserve the original value.

    When invoice_count is set, only that many rows are updated (via subquery
    with LIMIT). The rest stay with the original team_member.
    """
    inner_params: list[Any] = [row["week_start"], row["company_code"]]

    where_clauses = [
        "week_start = ?",
        "company_code = ?",
        "is_csv = 2",          # Only Envoy invoices (manual/CSV already have correct user_id)
    ]

    # Day filter
    if row["day_digits"] is not None:
        placeholders = ", ".join("?" * len(row["day_digits"]))
        where_clauses.append(
            f"strftime('%w', entry_date) IN ({placeholders})"
        )
        inner_params.extend(row["day_digits"])

    # Supplier filter
    if row["scope"] == "SUPPLIER":
        where_clauses.append("supplier_number = ?")
        inner_params.append(row["supplier_number"])

    # From-member filter
    if row["from_member"]:
        where_clauses.append("team_member = ?")
        inner_params.append(row["from_member"])

    where_sql = " AND ".join(where_clauses)

    invoice_count = row.get("invoice_count")
    if invoice_count:
        # Only update N invoices  the rest stay with the original member
        sql = (
            "UPDATE invoices\n"
            "SET original_team_member = COALESCE(original_team_member, team_member),\n"
            "    team_member = ?\n"
            "WHERE id IN (\n"
            f"    SELECT id FROM invoices WHERE {where_sql}\n"
            "    ORDER BY entry_date, id\n"
            "    LIMIT ?\n"
            ")"
        )
        params = [row["to_member"]] + inner_params + [invoice_count]
    else:
        # No limit  update ALL matching rows
        sql = (
            "UPDATE invoices\n"
            "SET original_team_member = COALESCE(original_team_member, team_member),\n"
            "    team_member = ?\n"
            f"WHERE {where_sql}"
        )
        params = [row["to_member"]] + inner_params
    return sql, params


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def apply_overrides(
    db_path: Path | None = None,
    overrides_path: Path | None = None,
    week: str | None = None,
    dry_run: bool = False,
) -> dict[str, Any]:
    """
    Apply team-member overrides from the weekly_overrides.xlsx file.

    Parameters
    ----------
    db_path:        Path to invoices.db. Defaults to scripts.paths.DB_PATH.
    overrides_path: Path to weekly_overrides.xlsx. Defaults to
                    scripts.paths.DATA_DIR / "weekly_overrides.xlsx".
    week:           If given, process only rows whose week_start matches.
    dry_run:        If True, print planned updates but do not execute them.

    Returns
    -------
    dict with keys: applied, rows_updated, skipped, errors
    """
    # Lazy imports so the module loads cleanly even if paths/config are absent
    from scripts.paths import DB_PATH, DATA_DIR  # noqa: PLC0415
    from scripts.config import ALL_MEMBERS, COMPANY_CODE_COUNTRY_MAP  # noqa: PLC0415  (incl. former members for historical rows)

    if db_path is None:
        db_path = DB_PATH
    if overrides_path is None:
        overrides_path = DATA_DIR / "weekly_overrides.xlsx"

    result: dict[str, Any] = {
        "applied": 0,
        "rows_updated": 0,
        "skipped": 0,
        "errors": [],
    }

    # --- Graceful degradation: missing file ---
    if not overrides_path.exists():
        logger.info("No overrides file found at %s  skipping.", overrides_path)
        return result

    # --- Graceful degradation: read Excel ---
    try:
        raw_rows = _read_overrides_file(overrides_path)
    except PermissionError:
        logger.warning(
            "Overrides file is locked (PermissionError): %s", overrides_path
        )
        return result
    except ImportError as exc:
        logger.error("Cannot import openpyxl: %s", exc)
        result["errors"].append(str(exc))
        return result
    except Exception as exc:  # noqa: BLE001
        logger.error("Failed to read overrides file: %s", exc)
        result["errors"].append(str(exc))
        return result

    if not raw_rows:
        logger.info("Overrides file is empty (no data rows after separator).")
        return result

    # --- Validate rows ---
    validated: list[dict[str, Any]] = []
    for idx, raw in enumerate(raw_rows, start=_DATA_START_ROW):
        # Week filter (applied before validation to avoid noise in logs)
        if week and raw["week_start"] and raw["week_start"] != week:
            continue

        cleaned = _validate_row(raw, ALL_MEMBERS, COMPANY_CODE_COUNTRY_MAP, idx)
        if cleaned is None:
            result["skipped"] += 1
        else:
            validated.append(cleaned)

    if not validated:
        logger.info("No valid override rows to apply (week filter: %s).", week)
        return result

    # --- Apply / dry-run ---
    with sqlite3.connect(db_path) as conn:
        migrate_schema(conn)

        for row in validated:
            sql, params = _build_update(row)

            if dry_run:
                # Count how many rows WOULD be updated
                count_sql = sql.replace(
                    "UPDATE invoices\n"
                    "SET original_team_member = COALESCE(original_team_member, team_member),\n"
                    "    team_member = ?\n",
                    "SELECT COUNT(*) FROM invoices\n",
                )
                # Remove to_member param (first param) since SELECT has no SET
                count_params = params[1:]
                count = conn.execute(count_sql, count_params).fetchone()[0]
                print(f"[DRY-RUN] Would update {count} row(s): {sql}")
                print(f"[DRY-RUN] Params: {params}\n")
                result["applied"] += 1
                result["rows_updated"] += count
                continue

            try:
                cursor = conn.execute(sql, params)
                updated = cursor.rowcount
                conn.commit()
                result["applied"] += 1
                result["rows_updated"] += updated
                logger.info(
                    "Override applied: week=%s cc=%s scope=%s to=%s  %d row(s) updated.",
                    row["week_start"],
                    row["company_code"],
                    row["scope"],
                    row["to_member"],
                    updated,
                )
            except sqlite3.Error as exc:
                msg = (
                    f"DB error on row week={row['week_start']} "
                    f"cc={row['company_code']}: {exc}"
                )
                logger.error(msg)
                result["errors"].append(msg)

    return result


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def _build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Apply team-member overrides from weekly_overrides.xlsx."
    )
    parser.add_argument(
        "--week",
        metavar="YYYY-MM-DD",
        default=None,
        help="Only process overrides for this week_start date.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        default=False,
        help="Print planned updates without executing them.",
    )
    parser.add_argument(
        "--db",
        metavar="PATH",
        default=None,
        help="Path to invoices.db (default: scripts.paths.DB_PATH).",
    )
    parser.add_argument(
        "--overrides",
        metavar="PATH",
        default=None,
        help="Path to weekly_overrides.xlsx (default: data/weekly_overrides.xlsx).",
    )
    return parser


if __name__ == "__main__":
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)-8s %(name)s  %(message)s",
    )

    parser = _build_arg_parser()
    args = parser.parse_args()

    summary = apply_overrides(
        db_path=Path(args.db) if args.db else None,
        overrides_path=Path(args.overrides) if args.overrides else None,
        week=args.week,
        dry_run=args.dry_run,
    )

    print(
        f"Done. applied={summary['applied']}  "
        f"rows_updated={summary['rows_updated']}  "
        f"skipped={summary['skipped']}  "
        f"errors={len(summary['errors'])}"
    )
    if summary["errors"]:
        for err in summary["errors"]:
            print(f"  ERROR: {err}")
