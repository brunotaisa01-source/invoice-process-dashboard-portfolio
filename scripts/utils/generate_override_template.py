"""
generate_override_template.py - Create the weekly overrides Excel template.

The TL fills this file each week to reassign team_member credit.
When the pipeline runs, overrides are applied to the SQLite database
before the dashboard is exported.

Usage:
    python -m scripts.utils.generate_override_template
    python -m scripts.utils.generate_override_template --output path/to/file.xlsx
"""
from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from scripts.config import COMPANY_CODE_COUNTRY_MAP, TEAM_MEMBERS
from scripts.paths import DATA_DIR


DEFAULT_OUTPUT = DATA_DIR / "weekly_overrides.xlsx"

COMPANY_CODES = sorted(COMPANY_CODE_COUNTRY_MAP.keys())
SCOPE_OPTIONS = ["ALL", "SUPPLIER"]
DAY_OPTIONS = ["FULL WEEK", "Mon", "Tue", "Wed", "Thu", "Fri"]

# --- Styles -----------------------------------------------------------

_HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

_EXAMPLE_FONT = Font(name="Calibri", italic=True, size=10, color="808080")
_EXAMPLE_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# --- Column definitions (name, width, header tooltip) ------------------
#     Col A=1, B=2, C=3, D=4, E=5, F=6, G=7, H=8, I=9

COLUMNS = [
    ("week_start",       16),   # A   dd/mm/yyyy (Friday = week start)
    ("day",              14),   # B   dropdown: FULL WEEK / Monday-Friday
    ("company_code",     14),   # C   dropdown
    ("scope",            12),   # D   dropdown: ALL / SUPPLIER
    ("supplier_number",  16),   # E   manual (only when scope=SUPPLIER)
    ("from_member",      20),   # F   dropdown (blank = all members)
    ("to_member",        20),   # G   dropdown (required)
    ("invoice_count",    14),   # H   manual
    ("days",             10),   # I   manual (0.5-5)
    ("notes",            35),   # J
]

_LAST_COL = "J"
_COL_DAY = 2
_COL_CC = 3
_COL_SCOPE = 4
_COL_SUPPLIER = 5
_COL_FROM = 6
_COL_TO = 7
_COL_INVOICES = 8
_COL_DAYS = 9

EXAMPLE_ROWS = [
    ("07/08/2026", "FULL WEEK", "SYN-CC-001", "ALL", "", "Synthetic Owner 001", "Synthetic Owner 002", 4, 2, "Synthetic weekly coverage"),
    ("07/08/2026", "Mon", "SYN-CC-002", "SUPPLIER", "SYN-SUP-002", "Synthetic Owner 002", "Synthetic Owner 001", 2, 1, "Synthetic single-day coverage"),
]


def _build_validation_sheet(wb: Workbook) -> None:
    """Create hidden _Validation sheet with lists for dropdowns."""
    ws = wb.create_sheet("_Validation")

    # Column A: team members
    ws["A1"] = "Team Members"
    ws["A1"].font = Font(bold=True)
    for i, name in enumerate(TEAM_MEMBERS, start=2):
        ws.cell(row=i, column=1, value=name)

    # Column B: company codes
    ws["B1"] = "Company Codes"
    ws["B1"].font = Font(bold=True)
    for i, cc in enumerate(COMPANY_CODES, start=2):
        ws.cell(row=i, column=2, value=cc)

    # Column C: country reference (for TL info)
    ws["C1"] = "Country"
    ws["C1"].font = Font(bold=True)
    for i, cc in enumerate(COMPANY_CODES, start=2):
        ws.cell(row=i, column=3, value=COMPANY_CODE_COUNTRY_MAP.get(cc, ""))

    # Column D: scope options
    ws["D1"] = "Scope"
    ws["D1"].font = Font(bold=True)
    for i, scope in enumerate(SCOPE_OPTIONS, start=2):
        ws.cell(row=i, column=4, value=scope)

    # Column E: day options
    ws["E1"] = "Days"
    ws["E1"].font = Font(bold=True)
    for i, day in enumerate(DAY_OPTIONS, start=2):
        ws.cell(row=i, column=5, value=day)

    ws.sheet_state = "hidden"


def _build_overrides_sheet(wb: Workbook) -> None:
    """Create the main Overrides sheet with headers, formatting, and validation."""
    ws = wb.create_sheet("Overrides", 0)

    # --- Title row ---
    ws.merge_cells(f"A1:{_LAST_COL}1")
    title_cell = ws["A1"]
    title_cell.value = "Weekly Team Member Overrides"
    title_cell.font = Font(name="Calibri", bold=True, size=14, color="2F5496")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # --- Instruction row ---
    ws.merge_cells(f"A2:{_LAST_COL}2")
    instr = ws["A2"]
    instr.value = (
        "Fill only rows where team member needs to change. "
        "Day: FULL WEEK or comma-separated days (Mon, Tue, Wed, Thu, Fri). "
        "Scope: ALL = entire company code, SUPPLIER = specific supplier. "
        "from_member blank = reassign ALL members' invoices."
    )
    instr.font = Font(name="Calibri", italic=True, size=9, color="666666")
    instr.alignment = Alignment(horizontal="left", wrap_text=True)
    ws.row_dimensions[2].height = 36

    # --- Headers (row 3) ---
    header_row = 3
    for col_idx, (name, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=name)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[header_row].height = 36

    # --- Example rows (greyed out) ---
    for row_offset, example in enumerate(EXAMPLE_ROWS):
        row_num = header_row + 1 + row_offset
        for col_idx, value in enumerate(example, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.font = _EXAMPLE_FONT
            cell.fill = _EXAMPLE_FILL
            cell.border = _THIN_BORDER
            if col_idx in (_COL_INVOICES, _COL_DAYS):
                cell.alignment = Alignment(horizontal="center")
            if col_idx == _COL_DAYS:
                cell.number_format = "0.0"

    # --- Separator row ---
    sep_row = header_row + len(EXAMPLE_ROWS) + 1
    ws.merge_cells(f"A{sep_row}:{_LAST_COL}{sep_row}")
    sep = ws.cell(row=sep_row, column=1)
    sep.value = "  Enter your overrides below  "
    sep.font = Font(name="Calibri", bold=True, size=10, color="2F5496")
    sep.alignment = Alignment(horizontal="center")
    sep.fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

    data_start = sep_row + 1
    data_end = 1000

    # --- Data validation: day (col B, free text for multi-select) ---
    dv_day = DataValidation(
        type="textLength",
        operator="greaterThanOrEqual",
        formula1="2",
        allow_blank=False,
    )
    dv_day.error = "Enter FULL WEEK or days: Mon, Tue, Wed, Thu, Fri"
    dv_day.errorTitle = "Invalid Day"
    dv_day.prompt = "FULL WEEK or comma-separated:\nMon, Tue, Wed, Thu, Fri"
    dv_day.promptTitle = "Day(s)"
    ws.add_data_validation(dv_day)
    dv_day.add(f"{get_column_letter(_COL_DAY)}{data_start}:{get_column_letter(_COL_DAY)}{data_end}")

    # --- Data validation: company_code (col C) ---
    n_cc = len(COMPANY_CODES)
    dv_cc = DataValidation(
        type="list",
        formula1=f"=_Validation!$B$2:$B${n_cc + 1}",
        allow_blank=False,
    )
    dv_cc.error = "Select a valid company code."
    dv_cc.errorTitle = "Invalid Company Code"
    dv_cc.prompt = "Synthetic company code (for example SYN-CC-001)"
    dv_cc.promptTitle = "Company Code"
    ws.add_data_validation(dv_cc)
    dv_cc.add(f"{get_column_letter(_COL_CC)}{data_start}:{get_column_letter(_COL_CC)}{data_end}")

    # --- Data validation: scope (col C) ---
    dv_scope = DataValidation(
        type="list",
        formula1=f"=_Validation!$D$2:$D${len(SCOPE_OPTIONS) + 1}",
        allow_blank=False,
    )
    dv_scope.error = "Select ALL or SUPPLIER."
    dv_scope.errorTitle = "Invalid Scope"
    dv_scope.prompt = "ALL = entire company code\nSUPPLIER = specific supplier"
    dv_scope.promptTitle = "Scope"
    ws.add_data_validation(dv_scope)
    dv_scope.add(f"{get_column_letter(_COL_SCOPE)}{data_start}:{get_column_letter(_COL_SCOPE)}{data_end}")

    # --- Data validation: from_member (col E, optional) ---
    n_members = len(TEAM_MEMBERS)
    dv_from = DataValidation(
        type="list",
        formula1=f"=_Validation!$A$2:$A${n_members + 1}",
        allow_blank=True,
    )
    dv_from.error = "Select a valid team member or leave blank for ALL."
    dv_from.errorTitle = "Invalid Member"
    dv_from.prompt = "Original member (blank = ALL members)"
    dv_from.promptTitle = "From Member"
    ws.add_data_validation(dv_from)
    dv_from.add(f"{get_column_letter(_COL_FROM)}{data_start}:{get_column_letter(_COL_FROM)}{data_end}")

    # --- Data validation: to_member (col F, required) ---
    dv_to = DataValidation(
        type="list",
        formula1=f"=_Validation!$A$2:$A${n_members + 1}",
        allow_blank=False,
    )
    dv_to.error = "Select a valid team member."
    dv_to.errorTitle = "Invalid Member"
    dv_to.prompt = "Who actually did the work"
    dv_to.promptTitle = "To Member"
    ws.add_data_validation(dv_to)
    dv_to.add(f"{get_column_letter(_COL_TO)}{data_start}:{get_column_letter(_COL_TO)}{data_end}")

    # --- Data validation: invoice_count (col G, whole >= 1) ---
    dv_inv = DataValidation(
        type="whole",
        operator="greaterThanOrEqual",
        formula1="1",
        allow_blank=True,
    )
    dv_inv.error = "Invoice count must be a whole number >= 1."
    dv_inv.errorTitle = "Invalid Count"
    dv_inv.prompt = "Optional - leave blank if unknown"
    dv_inv.promptTitle = "Invoice Count"
    ws.add_data_validation(dv_inv)
    dv_inv.add(f"{get_column_letter(_COL_INVOICES)}{data_start}:{get_column_letter(_COL_INVOICES)}{data_end}")

    # --- Data validation: days (col H, 0.5-5) ---
    dv_days = DataValidation(
        type="decimal",
        operator="between",
        formula1="0.5",
        formula2="5",
        allow_blank=True,
    )
    dv_days.error = "Days must be between 0.5 and 5."
    dv_days.errorTitle = "Invalid Days"
    dv_days.prompt = "Number of days (0.5 to 5)"
    dv_days.promptTitle = "Days Worked"
    ws.add_data_validation(dv_days)
    dv_days.add(f"{get_column_letter(_COL_DAYS)}{data_start}:{get_column_letter(_COL_DAYS)}{data_end}")

    # --- Date format for week_start (col A) ---
    for row_num in range(data_start, data_end + 1):
        ws.cell(row=row_num, column=1).number_format = "dd/mm/yyyy"

    # --- Freeze panes below header ---
    ws.freeze_panes = "A4"


def generate_template(output_path: Path | None = None) -> Path:
    """Generate the weekly_overrides.xlsx template.

    Args:
        output_path: Where to save. Defaults to data/weekly_overrides.xlsx

    Returns:
        Path to the generated file.
    """
    dest = output_path or DEFAULT_OUTPUT
    dest.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    _build_validation_sheet(wb)
    _build_overrides_sheet(wb)

    wb.save(str(dest))
    print(f"Template generated: {dest}")
    return dest


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Generate the weekly overrides Excel template."
    )
    parser.add_argument(
        "--output", "-o",
        type=Path,
        default=None,
        help=f"Output path (default: {DEFAULT_OUTPUT})",
    )
    args = parser.parse_args()
    generate_template(args.output)


if __name__ == "__main__":
    main()
